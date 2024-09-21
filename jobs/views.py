import requests
import os
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.utils import timezone
from .models import JobPost, JobApplication
from .forms import JobPostForm, JobApplicationForm, JobSearchForm, ResumeUploadForm
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.contrib import messages
import logging
from datetime import timedelta
from django.db.models import Q
import openai
import PyPDF2
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
import matplotlib.pyplot as plt
import io
import base64
import numpy as np
from django.views.generic import DetailView
from .utils import calculate_similarity, get_openai_analysis
import matplotlib
from botocore.exceptions import NoCredentialsError, ClientError
import boto3
from botocore.exceptions import NoCredentialsError, ClientError
import openpyxl
from openpyxl.utils import get_column_letter
from langdetect import detect
from googletrans import Translator
from payments.models import Order  
from django.urls import reverse
import json
import base64
import hashlib
from uuid import uuid4
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

PUBLIC_KEY = os.getenv('PUBLIC_KEY')
PRIVATE_KEY = os.getenv('PRIVATE_KEY')
EPOINT_API_URL = 'https://epoint.az/api/1/request'



# Initialize s3_client with Wasabi configuration
s3_client = boto3.client(
    's3',
    endpoint_url='https://s3.eu-central-2.wasabisys.com', 
    aws_access_key_id=os.getenv('R_ACCESS_KEY_ID'),
    aws_secret_access_key=os.getenv('R_SECRET_ACCESS_KEY')
)

matplotlib.use('Agg')
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
# Log to console
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)




# candidate views
def job_list(request):
    job_title = request.GET.get('job_title', '')
    company = request.GET.get('company', '')
    query = Q(deleted=False)

    # Add filters for job title and company with case-insensitive search using icontains
    if job_title:
        query &= Q(title__icontains=job_title)  # Case-insensitive search for job title
    if company:
        query &= Q(company__icontains=company)  # Case-insensitive search for company

    now = timezone.now()

    # Define time thresholds for scraped and non-scraped jobs
    scraped_threshold = now - timedelta(days=10)
    non_scraped_threshold = now - timedelta(days=15)

    # Retrieve non-scraped jobs
    non_scraped_jobs = JobPost.objects.filter(
        query,
        is_scraped=False,
        posted_at__gte=non_scraped_threshold
    ).order_by('-posted_at')

    # Retrieve scraped jobs
    scraped_jobs = JobPost.objects.filter(
        query,
        is_scraped=True,
        posted_at__gte=scraped_threshold
    ).order_by('-posted_at')

    # Combine non-scraped and scraped jobs while ensuring no duplicates
    combined_jobs = list(non_scraped_jobs) + list(scraped_jobs)

    # Use a dictionary to remove duplicates based on (title, company, apply_link)
    unique_jobs = {(job.title.lower(), job.company.lower(), job.apply_link): job for job in combined_jobs}.values()

    # Final pagination for the combined unique result
    final_paginator = Paginator(list(unique_jobs), 10)  # Show 10 jobs per page
    page = request.GET.get('page', 1)

    try:
        jobs_page = final_paginator.page(page)
    except PageNotAnInteger:
        jobs_page = final_paginator.page(1)
    except EmptyPage:
        jobs_page = final_paginator.page(final_paginator.num_pages)

    # Render the page with the jobs
    return render(request, 'jobs/job_list.html', {'jobs': jobs_page, 'job_title': job_title, 'company': company})

def upload_file_to_wasabi(file_name, bucket_name):
    try:
        # Check if the bucket exists
        s3_client.head_bucket(Bucket=bucket_name)
        logger.debug(f"Bucket '{bucket_name}' exists. Uploading file...")

        # Upload the file with public-read ACL
        s3_client.upload_file(file_name, bucket_name, file_name, ExtraArgs={'ACL': 'public-read'})
        logger.debug(f"File '{file_name}' uploaded successfully.")
        # Generate the file URL
        file_url = f"https://.s3.eu-central-2.wasabisys.com/{bucket_name}/resumes/{file_name}"
        logger.debug(f"File URL: {file_url}")
        return file_url
    except ClientError as e:
        if e.response['Error']['Code'] == '404':
            logger.error(f"Bucket '{bucket_name}' does not exist.")
        else:
            logger.error(f"Failed to upload file: {e}")
        return None
    except FileNotFoundError:
        logger.error("The file was not found.")
        return None
    except NoCredentialsError:
        logger.error("Credentials not available.")
        return None

def apply_job(request, job_id):
    job = get_object_or_404(JobPost, id=job_id)
    
    if request.method == 'POST':
        form = JobApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.job = job

            if 'resume' in request.FILES:
                resume = request.FILES['resume']
                file_name = f'resumes/{resume.name}'
                bucket_name = os.getenv('R_SPACES_NAME')

                try:
                    file_ext = resume.name.split('.')[-1].lower()
                    if file_ext == 'pdf':
                        cv_text = parse_pdf(resume)  # Parse the resume before uploading
                    else:
                        logger.error(f"Unsupported file format: {file_ext}")
                        messages.error(request, "Unsupported file format. Only PDF is supported.")
                        return redirect('apply_job', job_id=job.id)

                    job_description = job.description
                    similarity_score = calculate_similarity(cv_text, job_description)
                    application.match_score = similarity_score if similarity_score is not None else 0.0

                    with resume.open('rb') as resume_file:
                        s3_client.upload_fileobj(
                            resume_file, 
                            bucket_name, 
                            file_name, 
                            ExtraArgs={'ACL': 'public-read'}
                        )
                        application.resume = file_name

                except ClientError as e:
                    logger.error(f"Failed to upload resume: {e}")
                    messages.error(request, "Failed to upload resume. Please try again.")
                    return redirect('apply_job', job_id=job.id)
                except Exception as e:
                    logger.error(f"Error processing resume: {e}")
                    messages.error(request, "Failed to process resume. Please try again.")
                    return redirect('apply_job', job_id=job.id)

            application.save()
            return redirect('congrats')
    else:
        form = JobApplicationForm()
    
    return render(request, 'jobs/apply_job.html', {'form': form, 'job': job})


# hr views
@login_required
def post_job(request):
    if request.method == 'POST':
        form = JobPostForm(request.POST)
        if form.is_valid():
            job = form.save(commit=False)
            job.posted_by = request.user
            job.is_paid = False  # Mark job as not paid initially
            job.save()

            # Redirect to the payment process in the payments app
            return redirect('initiate_payment', job_id=job.id)  # Job ID will be passed to the payments app
    else:
        form = JobPostForm()

    return render(request, 'jobs/post_job.html', {'form': form})

@login_required
def edit_job(request, job_id):
    job = get_object_or_404(JobPost, id=job_id, posted_by=request.user, deleted=False)
    if request.method == 'POST':
        form = JobPostForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            return redirect('job_list')
    else:
        form = JobPostForm(instance=job)
    return render(request, 'jobs/edit_job.html', {'form': form, 'job': job})

@login_required
def delete_job(request, job_id):
    job = get_object_or_404(JobPost, id=job_id, posted_by=request.user)
    if request.method == 'POST':
        job.deleted = True
        job.save()
        return redirect('job_list')
    return render(request, 'jobs/confirm_delete.html', {'job': job})

@login_required
def hr_dashboard(request):
    if request.user.user_type != 'HR':
        return HttpResponseForbidden("You are not authorized to view this page.")

    # Search functionality
    search_query = request.GET.get('q', '')
    jobs = JobPost.objects.filter(posted_by=request.user, deleted=False)
    
    if search_query:
        jobs = jobs.filter(title__icontains=search_query)

    # Pagination setup
    jobs_page = request.GET.get('jobs_page', 1)
    jobs_paginator = Paginator(jobs, 5)
    try:
        jobs = jobs_paginator.page(jobs_page)
    except PageNotAnInteger:
        jobs = jobs_paginator.page(1)
    except EmptyPage:
        jobs = jobs_paginator.page(jobs_paginator.num_pages)

    return render(request, 'jobs/hr_dashboard.html', {'jobs': jobs, 'search_query': search_query})

@login_required
def hr_applicants(request, job_id):
    if request.user.user_type != 'HR':
        return HttpResponseForbidden("You are not authorized to view this page.")

    job = get_object_or_404(JobPost, id=job_id, posted_by=request.user)
    applications = JobApplication.objects.filter(job=job).exclude(full_name__isnull=True).order_by('-match_score', '-applied_at')

    applications_page = request.GET.get('applications_page', 1)
    applications_paginator = Paginator(applications, 30)
    try:
        applications = applications_paginator.page(applications_page)
    except PageNotAnInteger:
        applications = applications_paginator.page(1)
    except EmptyPage:
        applications = applications_paginator.page(applications_paginator.num_pages)

    return render(request, 'jobs/hr_applicants.html', {'applications': applications, 'job': job})

@login_required
def download_applicants_xlsx(request, job_id):
    if request.user.user_type != 'HR':
        return HttpResponseForbidden("You are not authorized to view this page.")

    job = get_object_or_404(JobPost, id=job_id, posted_by=request.user)
    applications = JobApplication.objects.filter(job=job).exclude(full_name__isnull=True).order_by('-applied_at')

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Applicants for {job.title}"

    # Define the headers
    headers = ['Full Name', 'Email', 'Phone Number', 'Applied At', 'Match Score', 'CV Download Link']
    ws.append(headers)

    # Write data to the sheet
    for application in applications:
        ws.append([
            application.full_name,
            application.email,
            application.phone,
            application.applied_at.strftime('%Y-%m-%d %H:%M'),
            application.match_score,
            f"{application.resume.url}" if application.resume else 'No CV Uploaded'
        ])

    # Set the width of the columns
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=applicants_{job_id}.xlsx'
    wb.save(response)
    return response

@login_required
def job_applicants(request, job_id):
    job = get_object_or_404(JobPost, id=job_id, posted_by=request.user, deleted=False)
    applications = JobApplication.objects.filter(job=job)
    return render(request, 'jobs/job_applicants.html', {'job': job, 'applications': applications})


# apply and create match score based on resume and job description
def parse_pdf(file):
    try:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)
        full_text = []
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            if text:
                full_text.append(text)
            else:
                raise ValueError(f"Unable to extract text from page {page_num + 1}")
        return '\n'.join(full_text)
    except PyPDF2.errors.PdfReadError as e:
        logger.error(f"Error reading PDF file: {e}")
        raise ValueError("The PDF file is unreadable or corrupted.")
    except Exception as e:
        logger.error(f"Unexpected error while parsing PDF: {e}")
        raise ValueError("An unexpected error occurred while processing the PDF.")

def translate_text(text, target_lang='en'):
    translator = Translator()
    try:
        translation = translator.translate(text, dest=target_lang)
        return translation.text
    except Exception as e:
        logger.error(f"Translation error: {e}")
        return None

def calculate_similarity(cv_text, job_text):
    # Detect languages of the CV and job description
    cv_lang = detect(cv_text)
    job_lang = detect(job_text)
    
    # If the CV or job description is not in English, translate them to English
    if cv_lang != 'en':
        cv_text = translate_text(cv_text, target_lang='en')
        if cv_text is None:
            return None  # Handle the case where translation fails
    if job_lang != 'en':
        job_text = translate_text(job_text, target_lang='en')
        if job_text is None:
            return None  # Handle the case where translation fails
    
    # Use the translated text for similarity calculation
    vectorizer = TfidfVectorizer().fit_transform([cv_text, job_text])
    vectors = vectorizer.toarray()
    return cosine_similarity(vectors)[0, 1]


# instructions and common views
def redirect_to_jobs(request):
    return redirect('job_list')

def congrats(request):
    return render(request, 'jobs/congrats.html')

def about(request):
    return render(request, 'jobs/about.html')

def robots_txt(request):
    lines = [
        "User-Agent: *",
        "Disallow: /admin/",
        "Sitemap: https://www.careerhorizon.llc/sitemap.xml"
    ]
    return HttpResponse("\n".join(lines), content_type="text/plain")
